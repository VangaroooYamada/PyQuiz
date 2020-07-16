import sys
import glob
import re
import time

import yaml
import openpyxl as px
import pyocr
import pyocr.builders
from PIL import Image
from tqdm import tqdm


# OCR Setting ***************
tools = pyocr.get_available_tools()
if len(tools) == 0:
    print("No OCR tool found")
    sys.exit(1)
tool = tools[0]     # Setting OCR tool
print("Will use tool '%s'" % (tool.get_name()))

langs = tool.get_available_languages()
print("Available languages: %s" % ", ".join(langs))
lang = langs[1]     # Setting Language for OCR (default 'jpn')
print("Will use lang '%s'\n" % (lang))


# Data for Excel (Quiz List) ***************
inames = list(glob.iglob('./imgs/*.png')) + list(glob.iglob('./imgs/*.jpg'))     # Image Path List
QLPath = './QuizList.xlsx'
quiz_list = px.load_workbook(QLPath)
with open('./trans_dict.yaml', encoding='utf-8') as yf:
    trans_dict = yaml.safe_load(yf)     # Dictionary for Translation txt


def time_counter(func):     # Decolator to count processing time
    def count_func(*args, **kwargs):
        start = time.time()
        i_count, q_count = func(*args, **kwargs)
        end = time.time() - start
        return i_count, q_count, end
    return count_func


def excel_safesaver(func):      # Decolator to check Excel's status
    def processing(*args, **kwargs):
        i_count, q_count = func(*args, **kwargs)
        for i in range(3):
            try:
                open('./QuizList.xlsx', 'r+')
            except IOError:
                print('*****************************************')
                print('Excel file is opened! Please close Excel.')
                print('({} times left)'.format(3 - i))
                print('*****************************************')
                _ = input('Close Excel and press any key.')
            else:
                break
        else:
            print('Processing Error!')
            quit(1)

        quiz_list.save('./QuizList.xlsx')
        return i_count, q_count
    return processing


@time_counter
@excel_safesaver
def image_processor(img_gen):   # method to write processed text to Excel
    ws = quiz_list.worksheets[-1]
    MR = quiz_list.worksheets[0].max_row
    BR = 2          # Brank Row's num
    i_count = 0     # Processed Image Counter
    q_count = 0     # Added Quiz Counter

    while BR < MR+1:
        if not ws['B' + str(BR)].value is None:
            BR += 1
            continue
        break

    pbar = tqdm(inames)

    for img in pbar:
        pbar.set_description('PROGRESS')

        txt = tool.image_to_string(
            Image.open(img),
            lang='jpn',
            builder=pyocr.builders.TextBuilder(tesseract_layout=3)
        )

        txt = txt.translate(str.maketrans(trans_dict)).replace('\n', '')
        for q in re.finditer(r'(正解率:\d+\%|Q\.)(.*?でしょう？)', txt):
            # print(q.groups()[1])
            if BR > MR:
                tmp_num = int(ws.title[:-1]) + 1000
                quiz_list.copy_worksheet(quiz_list.worksheets[0])
                ws = quiz_list.worksheets[-1]
                ws.title = str(tmp_num) + '-'
                BR = 2

            for question in q.split('Q.'):
                ws['B' + str(BR)].value = q.groups()[1]
                q_count += 1
                BR += 1

        i_count += 1

    return i_count, q_count


if __name__ == '__main__':

    i_count, q_count, end_time = image_processor(inames)

    if not i_count:
        print('No Image is processed.')
        print('Please put any image to "img" directory.')
    else:
        print('''
**********************
SUCCESS
{} Photos Processed
{} Quiz Added
Process Time: {}s
**********************
        '''.format(i_count, q_count, round(end_time, 3)))
